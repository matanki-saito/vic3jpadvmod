# -*- coding: utf-8 -*-
import os
import datetime
import openpyxl
from openpyxl.styles import Alignment
import re
from github import Github

label_en_map = {
	"イベント等テキスト": "event",
	"ゲームシステム用語": "term",
	"バグ": "term",
	"インターフェイス": "interface",
	"固有名詞": "proper noun",
	"固有名詞（バルク）": "proper nown (bulk)",
	"ツールチップ": "tooltip",
	"提案": "suggestion"
}

def create_xl():

	# トークンでgithubにアクセス
	g = Github(os.environ.get('GITHUB_TOKEN'))

	# リポジトリにアクセス
	repo = g.get_repo('matanki-saito/vic3jpadvmod')

	# excelブックの準備
	book = openpyxl.Workbook()
	book.remove(book.worksheets[-1])
	book.create_sheet(title='other')

	# 反復用の変数の準備/2行目と3列目
	ITR_FIRST_COL = 3
	i = ITR_FIRST_COL
	# 各issueに対して処理
	for _issue in repo.get_issues(state='closed'):
		issue = _issue

		# issueの本文を整形する
		# issueの本文を改行ごとに分ける
		lines = re.split('\r\n|\n', issue.body)
		body_continue = False
		headers = []
		bodies = []

		for line in lines:
			lfix = line.replace('```', '').replace("##", "").replace("（必須）", "").replace("（任意）", "")
			lfix = re.sub(r"!\[[^]]*]\(([^)]+)\)", r'=HYPERLINK("\1","screenshot")', lfix)
			# Excelに不要な行は削除
			if line in ['※「｀｀｀」は消さないでください', '※「```」は消さないでください', '＜ゲームシステム用語＞', '', '※画像はドラッグ＆ドロップで貼り付けられます。',
						'※できるだけ問題箇所の前後を含めて貼ってください。']:
				pass
			# 見出しを格納
			elif '##' in line:
				headers.append(lfix)
				body_continue = False
			# 本文を格納
			else:
				if len(bodies) == 0 or body_continue == False:
					bodies.append(lfix)
					bodies[-1] = bodies[-1]+'\n'
					body_continue = True
				elif body_continue:
					bodies[-1] = bodies[-1]+lfix+'\n'
					body_continue = True
			
		# issueに(タグ)がついていて、タグ名のシートがなければシートを作成
		if issue.labels:
			name = issue.labels[0].name
			if len(issue.labels) > 1:
				sub_label = issue.labels[1].name
				if sub_label == '提案':
					continue

			# 重複タグは起票済みをCloseしたものなので含まないようにする
			if name in ['重複', 'プログラム', '提案'] or name not in label_en_map.keys():
				continue

			e_name = label_en_map.get(name)

			if e_name in book.sheetnames:
				pass
			else:
				book.create_sheet(title=e_name)
			# タグ名のシートを選択
			sheet = book[e_name]

		# タグが付いていない場合'その他'のシートを選択
		else:
			sheet = book['other']

		# 出来立てのシートには一行目に見出しを付ける
		if sheet.max_row == 1:
			sheet.cell(row=1, column=1).value = 'number'
			sheet.cell(row=1, column=2).value = 'title'
			for h in headers:
				sheet.cell(row=1, column=sheet.max_column+1).value = h

		# issueの番号とタイトルを挿入する
		sheet.cell(row=sheet.max_row+1, column=1).value = \
			'=HYPERLINK("https://github.com/matanki-saito/vic3jpadvmod/issues/%d",%d)' % (issue.number, issue.number)
		sheet.cell(row=sheet.max_row, column=2).value = issue.title

		# issueの内容を挿入するのは選択したシートの最大行
		j = sheet.max_row

		# issueの内容を挿入する
		height_max = 1
		for b in bodies:
			sheet.cell(row=j, column=i).value = b
			# 行の高さを出すため本文の長さから行数を計算
			height = int(len(b) * 2 / 40)
			if height > height_max:
				height_max = height
			i = i+1
		sheet.row_dimensions[sheet.cell(row=j, column=i).row].height = height_max * 15
		i = ITR_FIRST_COL

	# 最後に全シートに対してスタイルを設定する
	for s in book:
		for c in range(2, s.max_column):
			s.column_dimensions[s.cell(row=1, column=c).column_letter].width = 40
			for r in range(2, s.max_row):
				s.cell(row=r, column=c).alignment = Alignment(horizontal='general', vertical='center', wrapText=True)

	# .xlsxファイルの保存先(例)：./issues/2022-10-30.xlsx
	xlname = './issues/'+str(datetime.date.today())+'.xlsx'
	book.save(xlname)


def main():
	create_xl()


if __name__ == "__main__":
	main()
